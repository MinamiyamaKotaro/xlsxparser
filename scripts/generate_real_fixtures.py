#!/usr/bin/env python3
"""Generates the "real Excel data" fixtures under tests/fixtures/{normal,
complex,error,load}/ using openpyxl, so xlsxparser is tested against files
a real, widely-used OOXML-writing tool actually produces — not just the
hand-authored minimal XML in tests/fixtures/{normal,error,complex,load,
security}.rs.

Two categories need more than "just call openpyxl and save":

- error/: no legitimate tool ever writes a syntactically broken or
  dangling-reference file on purpose, so these start from a genuine
  openpyxl-authored package and then mutate one specific part
  (`_mutate_zip_entries`) the same way a real-world failure would —
  truncating a part mid-write (disk full, interrupted upload), or breaking
  a single relationship ID — rather than being invalid from the start.
  `out_of_bounds_sst` additionally layers on a hand-built
  xl/sharedStrings.xml, because openpyxl itself never writes one (see
  `basic_types`'s doc comment) but real Microsoft Excel output almost
  always does, and the out-of-bounds scenario specifically needs one to
  reference out of bounds.

- load/: genuinely large workbooks, generated for real by openpyxl
  (`massive_dense_accounting`, `thousand_sheets`). `massive_sst` again
  layers a hand-built large xl/sharedStrings.xml on top for the same
  reason as `out_of_bounds_sst`. `many_images` layers 200 `<xdr:pic>`
  anchors on top the same way `embedded_image` does for a single one (Issue
  #71 follow-up), to check `parse_drawing`'s cost stays linear in image
  count.

- complex/embedded_image.xlsx: an image anchored via `<drawing r:id>`
  (Issue #65). openpyxl *can* insert images through its own API, but only
  by shelling out to Pillow to read the file, which this script's
  dependency footprint (`pip install openpyxl` alone) doesn't want to grow
  just for a fixture. `_add_drawing_with_image` instead splices in the four
  OOXML parts a real image insertion produces, the same layering technique
  `_add_shared_strings_part` already uses for a different missing-from-
  openpyxl part.

Kept in the repository (rather than run-once-and-discard) so these binary
fixtures can be regenerated if openpyxl's output shape ever changes.

Usage: python3 scripts/generate_real_fixtures.py
Requires: pip install openpyxl
"""

import datetime
import os
import re
import struct
import zipfile
import zlib

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
NORMAL_DIR = os.path.join(ROOT, "tests", "fixtures", "normal")
COMPLEX_DIR = os.path.join(ROOT, "tests", "fixtures", "complex")
ERROR_DIR = os.path.join(ROOT, "tests", "fixtures", "error")
LOAD_DIR = os.path.join(ROOT, "tests", "fixtures", "load")


# --- zip post-processing helpers (for error/ and the SST-layering cases) ---


def _read_zip_entries(path):
    with zipfile.ZipFile(path) as z:
        return {name: z.read(name) for name in z.namelist()}


def _write_zip_entries(path, entries):
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in entries.items():
            z.writestr(name, data)


def _mutate_zip_entries(path, mutate_fn):
    """Reads back the package `path` (already saved by openpyxl), lets
    `mutate_fn(entries: dict[str, bytes])` edit it in place, and rewrites
    the same path.
    """
    entries = _read_zip_entries(path)
    mutate_fn(entries)
    _write_zip_entries(path, entries)


def _two_cell_anchor_xml(with_hyperlink):
    """A `twoCellAnchor` spanning B2:E9 — an image explicitly sized to (and
    resized with) a multi-cell range, i.e. placed *across* several cells
    rather than confined to one.
    """
    hlink = '<a:hlinkClick r:id="rIdHyperlink"/>' if with_hyperlink else ""
    return (
        "<xdr:twoCellAnchor>"
        "<xdr:from><xdr:col>1</xdr:col><xdr:colOff>10000</xdr:colOff>"
        "<xdr:row>1</xdr:row><xdr:rowOff>20000</xdr:rowOff></xdr:from>"
        "<xdr:to><xdr:col>4</xdr:col><xdr:colOff>0</xdr:colOff>"
        "<xdr:row>8</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        "<xdr:pic>"
        "<xdr:nvPicPr>"
        f'<xdr:cNvPr id="2" name="Picture 1">{hlink}</xdr:cNvPr>'
        "<xdr:cNvPicPr/>"
        "</xdr:nvPicPr>"
        '<xdr:blipFill><a:blip r:embed="rIdEmbed"/></xdr:blipFill>'
        "<xdr:spPr/>"
        "</xdr:pic>"
        "<xdr:clientData/>"
        "</xdr:twoCellAnchor>"
    )


def _one_cell_anchor_xml():
    """A `oneCellAnchor` at C5, sized (`xdr:ext`) well under a default
    column width (609,600 EMU) / row height (190,500 EMU) — an image
    confined *within* a single cell, as opposed to `_two_cell_anchor_xml`'s
    multi-cell span. No hyperlink, to also cover that `hyperlink_r_id`/
    `PendingImage.hyperlink_r_id` genuinely stays `None` when `a:hlinkClick`
    is absent, not just when the drawing type differs.
    """
    return (
        "<xdr:oneCellAnchor>"
        "<xdr:from><xdr:col>2</xdr:col><xdr:colOff>5000</xdr:colOff>"
        "<xdr:row>4</xdr:row><xdr:rowOff>5000</xdr:rowOff></xdr:from>"
        '<xdr:ext cx="400000" cy="150000"/>'
        "<xdr:pic>"
        "<xdr:nvPicPr>"
        '<xdr:cNvPr id="2" name="Picture 1"/>'
        "<xdr:cNvPicPr/>"
        "</xdr:nvPicPr>"
        '<xdr:blipFill><a:blip r:embed="rIdEmbed"/></xdr:blipFill>'
        "<xdr:spPr/>"
        "</xdr:pic>"
        "<xdr:clientData/>"
        "</xdr:oneCellAnchor>"
    )


def _grouped_images_anchor_xml():
    """Two pictures grouped together via `<xdr:grpSp>` (Issue #67), using
    the same numeric conventions confirmed against real LibreOffice output:
    the outermost group's own `chOff`/`chExt` equal its `off`/`ext` (scale
    1), and both group- and pic-level `off`/`ext` are literal
    absolute-canvas EMU. The first pic carries no hyperlink; the second
    carries an `External` one, so the fixture also exercises per-pic
    hyperlink scoping surviving the group transform.
    """
    return (
        '<xdr:twoCellAnchor editAs="absolute">'
        "<xdr:from><xdr:col>1</xdr:col><xdr:colOff>267120</xdr:colOff>"
        "<xdr:row>4</xdr:row><xdr:rowOff>69840</xdr:rowOff></xdr:from>"
        "<xdr:to><xdr:col>3</xdr:col><xdr:colOff>441720</xdr:colOff>"
        "<xdr:row>7</xdr:row><xdr:rowOff>122040</xdr:rowOff></xdr:to>"
        "<xdr:grpSp>"
        '<xdr:nvGrpSpPr><xdr:cNvPr id="1" name=""/><xdr:cNvGrpSpPr/></xdr:nvGrpSpPr>'
        "<xdr:grpSpPr><a:xfrm>"
        '<a:off x="1080000" y="720000"/><a:ext cx="1800000" cy="540000"/>'
        '<a:chOff x="1080000" y="720000"/><a:chExt cx="1800000" cy="540000"/>'
        "</a:xfrm></xdr:grpSpPr>"
        "<xdr:pic>"
        '<xdr:nvPicPr><xdr:cNvPr id="2" name="Picture 1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rIdEmbed1"/></xdr:blipFill>'
        '<xdr:spPr><a:xfrm><a:off x="1080000" y="720000"/><a:ext cx="720000" cy="360000"/></a:xfrm></xdr:spPr>'
        "</xdr:pic>"
        "<xdr:pic>"
        '<xdr:nvPicPr><xdr:cNvPr id="3" name="Picture 2"><a:hlinkClick r:id="rIdHyperlink"/></xdr:cNvPr><xdr:cNvPicPr/></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rIdEmbed2"/></xdr:blipFill>'
        '<xdr:spPr><a:xfrm><a:off x="2160000" y="720000"/><a:ext cx="720000" cy="540000"/></a:xfrm></xdr:spPr>'
        "</xdr:pic>"
        "</xdr:grpSp>"
        "<xdr:clientData/>"
        "</xdr:twoCellAnchor>"
    )


def _tiny_png_bytes():
    """A minimal, genuinely-valid 1x1 transparent PNG (~70 bytes), built
    from stdlib `zlib`/`struct` alone — no Pillow dependency. Used in place
    of `sample_image.png` wherever a fixture needs *many* embedded media
    entries (`many_images`): xlsxparser never reads image bytes (Issue #65's
    stated scope stops at the anchor and the target path), so 200 real
    copies of the ~37KB `sample_image.png` would only bloat the repository
    for no verification benefit.
    """

    def chunk(tag, data):
        return (
            struct.pack(">I", len(data))
            + tag
            + data
            + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)
        )

    signature = b"\x89PNG\r\n\x1a\n"
    ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 6, 0, 0, 0)  # 1x1, 8-bit RGBA
    raw_scanline = b"\x00\x00\x00\x00\x00"  # filter byte + 1 RGBA pixel
    idat = zlib.compress(raw_scanline)
    return (
        signature
        + chunk(b"IHDR", ihdr)
        + chunk(b"IDAT", idat)
        + chunk(b"IEND", b"")
    )


def _many_images_anchor_xml(n):
    """`n` non-grouped `twoCellAnchor`s, one per row (image `i` spans row
    `i` to `i + 1` in column A), each with its own `rIdEmbed{i}` — a
    many-pictures-in-one-sheet scenario (Issue #71 follow-up: verifying
    parse_drawing's cost stays linear, not quadratic, in image count) as a
    real openpyxl-adjacent package rather than only a synthetic in-memory
    XML string.
    """
    anchors = []
    for i in range(n):
        anchors.append(
            "<xdr:twoCellAnchor>"
            f"<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{i}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{i + 1}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
            "<xdr:pic>"
            "<xdr:nvPicPr>"
            f'<xdr:cNvPr id="{i + 2}" name="Picture {i + 1}"/>'
            "<xdr:cNvPicPr/>"
            "</xdr:nvPicPr>"
            f'<xdr:blipFill><a:blip r:embed="rIdEmbed{i}"/></xdr:blipFill>'
            "<xdr:spPr/>"
            "</xdr:pic>"
            "<xdr:clientData/>"
            "</xdr:twoCellAnchor>"
        )
    return "".join(anchors)


def _add_drawing_with_many_images(entries, n):
    """Like `_add_drawing_with_image`, but registers `n` embedded media
    relationships (`rIdEmbed0`..`rIdEmbed{n-1}`) for `_many_images_anchor_xml`'s
    `n` independent pictures. All `n` media entries reuse the same
    `_tiny_png_bytes()` (see its docstring for why a real photo isn't used
    here, unlike `_add_drawing_with_grouped_images`).
    """
    image_bytes = _tiny_png_bytes()
    for i in range(n):
        entries[f"xl/media/image{i}.png"] = image_bytes

    entries["xl/drawings/drawing1.xml"] = (
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        + _many_images_anchor_xml(n)
        + "</xdr:wsDr>"
    ).encode("utf-8")

    rels = [
        f'<Relationship Id="rIdEmbed{i}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        f'Target="../media/image{i}.png"/>'
        for i in range(n)
    ]
    entries["xl/drawings/_rels/drawing1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(rels)
        + "</Relationships>"
    ).encode("utf-8")

    entries["xl/worksheets/_rels/sheet1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdDrawing" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert "<drawing " not in sheet_xml
    sheet_xml = sheet_xml.replace(
        "</worksheet>", '<drawing r:id="rIdDrawing"/></worksheet>'
    )
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")

    content_types = entries["[Content_Types].xml"].decode("utf-8")
    content_types = content_types.replace(
        "</Types>",
        '<Default Extension="png" ContentType="image/png"/>'
        '<Override PartName="/xl/drawings/drawing1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        "</Types>",
    )
    entries["[Content_Types].xml"] = content_types.encode("utf-8")


def _add_drawing_with_grouped_images(entries, image_path):
    """Like `_add_drawing_with_image`, but splices in two embedded media
    relationships (`rIdEmbed1`/`rIdEmbed2`) plus one hyperlink
    (`rIdHyperlink`, only on the second pic) for `_grouped_images_anchor_xml`'s
    two-picture `<xdr:grpSp>`. Both media entries reuse the same
    `sample_image.png` bytes — xlsxparser never reads them either way.
    """
    with open(image_path, "rb") as f:
        image_bytes = f.read()
    entries["xl/media/image1.png"] = image_bytes
    entries["xl/media/image2.png"] = image_bytes

    entries["xl/drawings/drawing1.xml"] = (
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        + _grouped_images_anchor_xml()
        + "</xdr:wsDr>"
    ).encode("utf-8")

    entries["xl/drawings/_rels/drawing1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdEmbed1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        'Target="../media/image1.png"/>'
        '<Relationship Id="rIdEmbed2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        'Target="../media/image2.png"/>'
        '<Relationship Id="rIdHyperlink" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
        'Target="https://example.com/second-logo" TargetMode="External"/>'
        "</Relationships>"
    ).encode("utf-8")

    entries["xl/worksheets/_rels/sheet1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdDrawing" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert "<drawing " not in sheet_xml
    sheet_xml = sheet_xml.replace(
        "</worksheet>", '<drawing r:id="rIdDrawing"/></worksheet>'
    )
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")

    content_types = entries["[Content_Types].xml"].decode("utf-8")
    content_types = content_types.replace(
        "</Types>",
        '<Default Extension="png" ContentType="image/png"/>'
        '<Override PartName="/xl/drawings/drawing1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        "</Types>",
    )
    entries["[Content_Types].xml"] = content_types.encode("utf-8")


def _add_drawing_with_image(entries, image_path, anchor_xml, with_hyperlink):
    """Splices a `<drawing>`-anchored image onto sheet1, the same way
    `_add_shared_strings_part` splices in a part openpyxl's high-level API
    doesn't write directly — openpyxl *can* insert images via
    `openpyxl.drawing.image.Image`, but only by shelling out to Pillow to
    read the file, which this repository doesn't want as a fixture-generation
    dependency. Hand-building the four OOXML parts a real image insertion
    produces (`xl/media/image1.png`, `xl/drawings/drawing1.xml` — its
    anchor supplied by the caller as `anchor_xml`, either
    `_two_cell_anchor_xml` or `_one_cell_anchor_xml` — plus
    `xl/drawings/_rels/drawing1.xml.rels` and
    `xl/worksheets/_rels/sheet1.xml.rels`) sidesteps that while still
    exercising the genuine `<drawing r:id="...">` cross-reference a real
    writer emits (Issue #65). `drawing1.xml.rels`'s hyperlink relationship is
    only registered when `with_hyperlink` is set — an unreferenced
    relationship would still resolve fine, but omitting it keeps the "no
    hyperlink" fixture from implying one exists that nothing in
    `drawing1.xml` actually points at.
    """
    with open(image_path, "rb") as f:
        entries["xl/media/image1.png"] = f.read()

    entries["xl/drawings/drawing1.xml"] = (
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        + anchor_xml
        + "</xdr:wsDr>"
    ).encode("utf-8")

    hyperlink_rel = (
        '<Relationship Id="rIdHyperlink" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
        'Target="https://example.com/sample-image" TargetMode="External"/>'
        if with_hyperlink
        else ""
    )
    entries["xl/drawings/_rels/drawing1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdEmbed" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        'Target="../media/image1.png"/>'
        + hyperlink_rel
        + "</Relationships>"
    ).encode("utf-8")

    entries["xl/worksheets/_rels/sheet1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdDrawing" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert "<drawing " not in sheet_xml
    sheet_xml = sheet_xml.replace(
        "</worksheet>", '<drawing r:id="rIdDrawing"/></worksheet>'
    )
    entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")

    content_types = entries["[Content_Types].xml"].decode("utf-8")
    content_types = content_types.replace(
        "</Types>",
        '<Default Extension="png" ContentType="image/png"/>'
        '<Override PartName="/xl/drawings/drawing1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        "</Types>",
    )
    entries["[Content_Types].xml"] = content_types.encode("utf-8")


def _add_shared_strings_part(entries, strings):
    """Adds a hand-built, genuine-shaped xl/sharedStrings.xml to `entries`
    and registers its relationship in xl/_rels/workbook.xml.rels — the part
    openpyxl itself never writes (see module docstring) but real Microsoft
    Excel output does.
    """
    sst_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{len(strings)}" uniqueCount="{len(strings)}">'
        + "".join(f"<si><t>{s}</t></si>" for s in strings)
        + "</sst>"
    )
    entries["xl/sharedStrings.xml"] = sst_xml.encode("utf-8")

    rels_xml = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rels_xml = rels_xml.replace(
        "</Relationships>",
        '<Relationship Id="rIdSharedStrings" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
        'Target="sharedStrings.xml"/></Relationships>',
    )
    entries["xl/_rels/workbook.xml.rels"] = rels_xml.encode("utf-8")


# --- normal/ ---


def basic_types():
    """Every CellValue variant in one row, written by openpyxl rather than
    hand-authored XML. Notably, openpyxl writes text cells as
    t="inlineStr" (it dropped writing xl/sharedStrings.xml entirely in
    recent versions) — itself a real-world instance of the
    "third-party tool that emits inline strings" scenario
    tests/fixtures/normal.rs::inline_strings documents.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "日本語Text"
    ws["B1"] = 42
    ws["C1"] = 19.99
    ws["D1"] = datetime.date(2023, 6, 15)
    ws["D1"].number_format = "yyyy-mm-dd"
    ws["E1"] = True
    ws["F1"] = False
    # openpyxl has no direct "write an error value" API (error codes only
    # ever arise as cached formula results in a real workbook); set the
    # cell's data_type directly to produce a genuine t="e" cell.
    ws["G1"] = "#N/A"
    ws["G1"].data_type = "e"
    path = os.path.join(NORMAL_DIR, "basic_types.xlsx")
    wb.save(path)
    return path


# --- complex/ ---


def houganshi_merged():
    """A merged A1:C3 region, the "grid-paper Excel" shape from Issue #28."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "houganshi"
    ws.merge_cells("A1:C3")
    path = os.path.join(COMPLEX_DIR, "houganshi_merged.xlsx")
    wb.save(path)
    return path


def cell_hyperlinks():
    """Cell hyperlinks (Issue #95): an external URL with a tooltip (A2,
    the common case — openpyxl's `cell.hyperlink = "..."` API), and a
    `location`-only internal jump with no `r:id` at all (A3, via
    `openpyxl.worksheet.hyperlink.Hyperlink`). Real openpyxl output
    declares `xmlns:r` inline on the `<hyperlink>` element itself rather
    than on the `<worksheet>` root — a real-world XML shape none of
    xlsxparser's hand-authored fixtures happen to exercise, since they
    always declare `xmlns:r` at the root instead.
    """
    from openpyxl.worksheet.hyperlink import Hyperlink

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "click me"
    ws["A2"].hyperlink = "https://example.com/"
    ws["A2"].hyperlink.tooltip = "Visit example"
    ws["A3"] = "jump"
    ws["A3"].hyperlink = Hyperlink(ref="A3", location="Sheet1!A1")
    path = os.path.join(COMPLEX_DIR, "cell_hyperlinks.xlsx")
    wb.save(path)
    return path


def multi_sheet_states():
    """Visible / hidden / veryHidden / empty sheets in one workbook."""
    wb = openpyxl.Workbook()
    visible = wb.active
    visible.title = "Visible"
    visible["A1"] = 1

    hidden = wb.create_sheet("Hidden")
    hidden["A1"] = 1
    hidden.sheet_state = "hidden"

    very_hidden = wb.create_sheet("VeryHidden")
    very_hidden["A1"] = 1
    very_hidden.sheet_state = "veryHidden"

    # No cells written at all.
    wb.create_sheet("Empty")

    path = os.path.join(COMPLEX_DIR, "multi_sheet_states.xlsx")
    wb.save(path)
    return path


def embedded_image():
    """A picture anchored to B2:E9 via `twoCellAnchor` (Issue #65) — placed
    *across* a multi-cell range rather than confined to one, carrying both
    an embedded media relationship and the image's own (External)
    hyperlink, the two relationship kinds `pipeline.rs`'s Phase 3.5 resolves
    against `drawing1.xml.rels`. Uses the checked-in
    `scripts/fixtures_assets/sample_image.png` as the embedded media's real
    bytes, even though xlsxparser itself never reads them (Issue #65's
    stated scope stops at the anchor position and target *path*). Contrast
    `embedded_image_one_cell`, the single-cell-confined counterpart.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "logo"
    path = os.path.join(COMPLEX_DIR, "embedded_image.xlsx")
    wb.save(path)

    image_path = os.path.join(ROOT, "scripts", "fixtures_assets", "sample_image.png")
    _mutate_zip_entries(
        path,
        lambda entries: _add_drawing_with_image(
            entries, image_path, _two_cell_anchor_xml(with_hyperlink=True), with_hyperlink=True
        ),
    )
    return path


def embedded_image_one_cell():
    """A picture anchored at C5 via `oneCellAnchor` (Issue #65), sized well
    under a default cell's dimensions — confined *within* a single cell,
    unlike `embedded_image`'s multi-cell `twoCellAnchor` span. Has no
    hyperlink, so `Image::hyperlink` should resolve to `None`.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "icon"
    path = os.path.join(COMPLEX_DIR, "embedded_image_one_cell.xlsx")
    wb.save(path)

    image_path = os.path.join(ROOT, "scripts", "fixtures_assets", "sample_image.png")
    _mutate_zip_entries(
        path,
        lambda entries: _add_drawing_with_image(
            entries, image_path, _one_cell_anchor_xml(), with_hyperlink=False
        ),
    )
    return path


def grouped_images():
    """Two pictures grouped via `<xdr:grpSp>` (Issue #67), mirroring
    `embedded_image`/`embedded_image_one_cell` but for the group-transform
    resolution path. See `_grouped_images_anchor_xml`'s docstring for the
    numeric layout.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "logos"
    path = os.path.join(COMPLEX_DIR, "grouped_images.xlsx")
    wb.save(path)

    image_path = os.path.join(ROOT, "scripts", "fixtures_assets", "sample_image.png")
    _mutate_zip_entries(
        path,
        lambda entries: _add_drawing_with_grouped_images(entries, image_path),
    )
    return path


def styled_fill_color():
    """Three cells with a `solid` pattern fill (Issue #75): A1 an RGB
    direct-specification fill, A2 a theme+tint fill, A3 a plain (no-fill)
    cell for contrast. Unlike `embedded_image`, openpyxl's own
    `PatternFill`/`Color` API covers this natively — no manual OOXML
    part-splicing is needed.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "rgb fill"
    ws["A1"].fill = PatternFill(
        start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
    )
    ws["A2"] = "theme fill"
    ws["A2"].fill = PatternFill(
        fgColor=Color(theme=4, tint=-0.25), fill_type="solid"
    )
    ws["A3"] = "no fill"
    path = os.path.join(COMPLEX_DIR, "styled_fill_color.xlsx")
    wb.save(path)
    return path


def extreme_sparse():
    """A1 and Excel's absolute bottom-right corner, XFD1048576, populated —
    nothing in between.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=1)
    ws.cell(row=1_048_576, column=16_384, value=2)
    path = os.path.join(COMPLEX_DIR, "extreme_sparse.xlsx")
    wb.save(path)
    return path


# --- error/ ---


def corrupted_xml():
    """A genuine openpyxl-authored worksheet part, truncated partway
    through — simulating a real-world interrupted write (disk full, killed
    process, a network upload cut short) rather than a file that was
    invalid XML from the start.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "will be truncated"
    ws["B1"] = 123
    path = os.path.join(ERROR_DIR, "corrupted_xml.xlsx")
    wb.save(path)

    def mutate(entries):
        sheet_xml = entries["xl/worksheets/sheet1.xml"]
        entries["xl/worksheets/sheet1.xml"] = sheet_xml[: len(sheet_xml) * 2 // 3]

    _mutate_zip_entries(path, mutate)
    return path


def missing_relations():
    """A genuine workbook.xml.rels with the worksheet's relationship Id
    changed, so workbook.xml's <sheet r:id="rId1"> no longer resolves —
    simulating real-world rels corruption (e.g. a re-zipping tool that
    renumbers relationship IDs inconsistently with the parts that
    reference them).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    path = os.path.join(ERROR_DIR, "missing_relations.xlsx")
    wb.save(path)

    def mutate(entries):
        rels_xml = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
        # openpyxl assigns the worksheet relationship Id="rId1" for a
        # single-sheet workbook; workbook.xml's <sheet> still references
        # "rId1" after this, so the reference now dangles.
        broken = rels_xml.replace('Id="rId1"', 'Id="rIdRenumbered"', 1)
        assert broken != rels_xml, "expected to find Id=\"rId1\" to break"
        entries["xl/_rels/workbook.xml.rels"] = broken.encode("utf-8")

    _mutate_zip_entries(path, mutate)
    return path


def invalid_merge_ref():
    """A genuine merged region (A1:C3), with only the <mergeCell> tag's
    `ref` reversed to C3:A1 after the fact — <dimension ref="A1:C3"/>
    (unrelated, and not read by xlsxparser at all) is deliberately left
    untouched to keep the mutation minimal and realistic.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "merged"
    ws.merge_cells("A1:C3")
    path = os.path.join(ERROR_DIR, "invalid_merge_ref.xlsx")
    wb.save(path)

    def mutate(entries):
        sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
        corrupted, count = re.subn(
            r'(<mergeCell ref=")A1:C3(")', r"\1C3:A1\2", sheet_xml
        )
        assert count == 1, "expected exactly one <mergeCell ref=\"A1:C3\"/>"
        entries["xl/worksheets/sheet1.xml"] = corrupted.encode("utf-8")

    _mutate_zip_entries(path, mutate)
    return path


def out_of_bounds_sst():
    """A genuine openpyxl-authored package with a hand-built, real-shaped
    xl/sharedStrings.xml layered on (1 entry), and A1 rewritten to
    reference shared-string index 99999 — out of bounds.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "placeholder"
    path = os.path.join(ERROR_DIR, "out_of_bounds_sst.xlsx")
    wb.save(path)

    def mutate(entries):
        _add_shared_strings_part(entries, ["only one entry"])
        sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
        rewritten, count = re.subn(
            r'<c r="A1"[^>]*>.*?</c>',
            '<c r="A1" t="s"><v>99999</v></c>',
            sheet_xml,
            count=1,
            flags=re.DOTALL,
        )
        assert count == 1, "expected exactly one <c r=\"A1\"> to rewrite"
        entries["xl/worksheets/sheet1.xml"] = rewritten.encode("utf-8")

    _mutate_zip_entries(path, mutate)
    return path


# --- load/ ---


def massive_dense_accounting():
    """A genuinely-generated dense 10,000-row x 30-column sheet (300,000
    cells) — no post-processing, this is exactly what openpyxl writes for
    that much data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    for row in range(1, 10_001):
        for col in range(1, 31):
            ws.cell(row=row, column=col, value=row * 100 + col)
    path = os.path.join(LOAD_DIR, "massive_dense_accounting.xlsx")
    wb.save(path)
    return path


def thousand_sheets():
    """1,000 genuinely-generated sheets, each holding a single cell."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for i in range(1, 1001):
        ws = wb.create_sheet(f"Sheet{i}")
        ws["A1"] = i
    path = os.path.join(LOAD_DIR, "thousand_sheets.xlsx")
    wb.save(path)
    return path


def massive_sst():
    """A genuine openpyxl-authored package with a hand-built 50,000-entry
    xl/sharedStrings.xml layered on, and 3 cells referencing the first, a
    middle, and the last index.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "placeholder-a"
    ws["B1"] = "placeholder-b"
    ws["C1"] = "placeholder-c"
    path = os.path.join(LOAD_DIR, "massive_sst.xlsx")
    wb.save(path)

    string_count = 50_000

    def mutate(entries):
        strings = [f"unique-string-{i}" for i in range(string_count)]
        _add_shared_strings_part(entries, strings)
        sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
        for ref, index in (
            ("A1", 0),
            ("B1", string_count // 2),
            ("C1", string_count - 1),
        ):
            sheet_xml, count = re.subn(
                rf'<c r="{ref}"[^>]*>.*?</c>',
                f'<c r="{ref}" t="s"><v>{index}</v></c>',
                sheet_xml,
                count=1,
                flags=re.DOTALL,
            )
            assert count == 1, f"expected exactly one <c r=\"{ref}\"> to rewrite"
        entries["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")

    _mutate_zip_entries(path, mutate)
    return path


def many_images():
    """200 non-grouped pictures, one per row (Issue #71 follow-up): a
    genuine `.xlsx` package to make the "does parse_drawing's cost stay
    linear (not quadratic) in image count" claim a permanent, checked-in
    regression test rather than a throwaway benchmark. 200 keeps the
    checked-in fixture and CI runtime small while still being far larger
    than any handful-of-images fixture elsewhere in this repository.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "many images"
    path = os.path.join(LOAD_DIR, "many_images.xlsx")
    wb.save(path)

    _mutate_zip_entries(
        path,
        lambda entries: _add_drawing_with_many_images(entries, 200),
    )
    return path


def main():
    for directory in (NORMAL_DIR, COMPLEX_DIR, ERROR_DIR, LOAD_DIR):
        os.makedirs(directory, exist_ok=True)
    for fn in (
        basic_types,
        houganshi_merged,
        cell_hyperlinks,
        multi_sheet_states,
        embedded_image,
        embedded_image_one_cell,
        grouped_images,
        styled_fill_color,
        extreme_sparse,
        corrupted_xml,
        missing_relations,
        invalid_merge_ref,
        out_of_bounds_sst,
        massive_dense_accounting,
        thousand_sheets,
        massive_sst,
        many_images,
    ):
        path = fn()
        print(f"wrote {os.path.relpath(path, ROOT)}")


if __name__ == "__main__":
    main()
